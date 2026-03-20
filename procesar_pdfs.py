"""
=============================================================
  PROCESADOR MASIVO DE PDFs → EXCEL
  Maneja PDFs digitales y escaneados (mixtos)
  Autor: generado con Claude
=============================================================

INSTALACIÓN (ejecutar una sola vez en terminal):
    pip install pdfplumber pypdf pdf2image pytesseract openpyxl tqdm pillow

ADEMÁS necesitas instalar Tesseract OCR (para PDFs escaneados):
    Windows: https://github.com/UB-Mannheim/tesseract/wiki
             Descargar e instalar el instalador .exe
             Asegúrate de marcar "Spanish" durante la instalación

    Mac:     brew install tesseract tesseract-lang

    Linux:   sudo apt install tesseract-ocr tesseract-ocr-spa

USO:
    1. Pon todos tus PDFs en una carpeta (ej: C:/mis_pdfs/)
    2. Cambia la variable CARPETA_PDFS abajo
    3. Ejecuta: python procesar_pdfs.py
    4. El resultado será: resultado_consolidado.xlsx
=============================================================
"""

import os
import re
import warnings
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tqdm import tqdm
from datetime import datetime

warnings.filterwarnings("ignore")

# ============================================================
# ⚙️  CONFIGURACIÓN — EDITA SOLO ESTA SECCIÓN
# ============================================================

CARPETA_PDFS = r"C:\Users\nicol\Desktop\Analisis del mapa python\pdf\3.9.2 - EJECUCION MANTENIMIENTO MECANICO PUERTO"          # <- Cambia esta ruta a tu carpeta
ARCHIVO_SALIDA = "resultado_consolidado.xlsx"
IDIOMA_OCR = "spa"                      # "spa" para español, "eng" para inglés
MINIMO_CHARS_TEXTO = 50                 # Si un PDF tiene menos chars, se trata como escaneado

# Ruta a Tesseract (solo Windows — ajusta si instalaste en otro lugar)
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ============================================================


def extraer_texto_digital(ruta_pdf):
    """Extrae texto de PDFs digitales usando pdfplumber."""
    texto_total = []
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            for i, pagina in enumerate(pdf.pages):
                texto = pagina.extract_text()
                if texto:
                    texto_total.append(f"[Página {i+1}]\n{texto.strip()}")
    except Exception as e:
        return None, str(e)
    return "\n\n".join(texto_total), None


def extraer_texto_ocr(ruta_pdf):
    """Extrae texto de PDFs escaneados usando OCR (Tesseract)."""
    texto_total = []
    try:
        imagenes = convert_from_path(ruta_pdf, dpi=200)
        for i, imagen in enumerate(imagenes):
            texto = pytesseract.image_to_string(imagen, lang=IDIOMA_OCR)
            if texto.strip():
                texto_total.append(f"[Página {i+1}]\n{texto.strip()}")
    except Exception as e:
        return None, str(e)
    return "\n\n".join(texto_total), None


def detectar_num_paginas(ruta_pdf):
    """Cuenta las páginas del PDF."""
    try:
        reader = PdfReader(ruta_pdf)
        return len(reader.pages)
    except:
        return "?"


def es_escaneado(texto):
    """Determina si un PDF necesita OCR basándose en el texto extraído."""
    if texto is None:
        return True
    return len(texto.strip()) < MINIMO_CHARS_TEXTO


def procesar_pdf(ruta_pdf):
    """Procesa un PDF y retorna texto + metadata."""
    nombre_archivo = os.path.basename(ruta_pdf)
    num_paginas = detectar_num_paginas(ruta_pdf)

    # Intentar extracción digital primero
    texto, error = extraer_texto_digital(ruta_pdf)

    if es_escaneado(texto):
        # Fallback a OCR para PDFs escaneados
        metodo = "OCR (Escaneado)"
        texto_ocr, error_ocr = extraer_texto_ocr(ruta_pdf)
        if texto_ocr and len(texto_ocr.strip()) > MINIMO_CHARS_TEXTO:
            texto = texto_ocr
            error = None
        elif error_ocr:
            error = error_ocr
    else:
        metodo = "Digital"

    estado = "✅ OK" if texto and len(texto.strip()) > 10 else "⚠️ Sin texto"
    if error:
        estado = f"❌ Error: {error[:80]}"

    return {
        "archivo": nombre_archivo,
        "ruta_completa": ruta_pdf,
        "paginas": num_paginas,
        "metodo": metodo,
        "estado": estado,
        "texto": (texto or "").strip(),
        "num_chars": len((texto or "").strip()),
    }


def crear_excel(resultados, archivo_salida):
    """Genera el Excel consolidado con formato."""
    wb = Workbook()

    # ── Hoja 1: RESUMEN ─────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    # Encabezados
    encabezados = ["#", "Archivo PDF", "Páginas", "Método", "Estado", "Caracteres"]
    ws_resumen.append(encabezados)

    # Estilo encabezado
    color_header = PatternFill("solid", fgColor="1F4E79")
    for celda in ws_resumen[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = color_header
        celda.alignment = Alignment(horizontal="center")

    # Datos
    ok = sum(1 for r in resultados if "✅" in r["estado"])
    ocr = sum(1 for r in resultados if "OCR" in r["metodo"])
    errores = sum(1 for r in resultados if "❌" in r["estado"])

    for i, r in enumerate(resultados, 1):
        fila = [i, r["archivo"], r["paginas"], r["metodo"], r["estado"], r["num_chars"]]
        ws_resumen.append(fila)
        # Color por estado
        fila_ws = ws_resumen[i + 1]
        if "❌" in r["estado"]:
            for c in fila_ws:
                c.fill = PatternFill("solid", fgColor="FFDEDE")
        elif "⚠️" in r["estado"]:
            for c in fila_ws:
                c.fill = PatternFill("solid", fgColor="FFF3CD")

    # Ajustar anchos
    ws_resumen.column_dimensions["B"].width = 45
    ws_resumen.column_dimensions["D"].width = 20
    ws_resumen.column_dimensions["E"].width = 35

    # ── Hoja 2: TEXTOS COMPLETOS ─────────────────────────────
    ws_textos = wb.create_sheet("Textos Completos")
    enc2 = ["#", "Archivo PDF", "Páginas", "Método", "Texto Extraído"]
    ws_textos.append(enc2)

    for celda in ws_textos[1]:
        celda.font = Font(bold=True, color="FFFFFF", size=11)
        celda.fill = PatternFill("solid", fgColor="1F4E79")
        celda.alignment = Alignment(horizontal="center")

    for i, r in enumerate(resultados, 1):
        texto_truncado = r["texto"][:32000] if len(r["texto"]) > 32000 else r["texto"]
        ws_textos.append([i, r["archivo"], r["paginas"], r["metodo"], texto_truncado])
        ws_textos[i + 1][4].alignment = Alignment(wrap_text=True, vertical="top")

    ws_textos.column_dimensions["B"].width = 45
    ws_textos.column_dimensions["D"].width = 20
    ws_textos.column_dimensions["E"].width = 80
    ws_textos.row_dimensions[1].height = 20

    # ── Hoja 3: ESTADÍSTICAS ─────────────────────────────────
    ws_stats = wb.create_sheet("Estadísticas")
    ws_stats["A1"] = "RESUMEN DE PROCESAMIENTO"
    ws_stats["A1"].font = Font(bold=True, size=14, color="1F4E79")

    stats = [
        ["", ""],
        ["📅 Fecha procesamiento", datetime.now().strftime("%d/%m/%Y %H:%M")],
        ["📁 Carpeta origen", CARPETA_PDFS],
        ["", ""],
        ["📊 Total PDFs procesados", len(resultados)],
        ["✅ Procesados correctamente", ok],
        ["🔍 Procesados con OCR", ocr],
        ["❌ Con errores", errores],
        ["", ""],
        ["📄 Total páginas procesadas", sum(r["paginas"] for r in resultados if isinstance(r["paginas"], int))],
        ["🔤 Total caracteres extraídos", sum(r["num_chars"] for r in resultados)],
    ]
    for fila in stats:
        ws_stats.append(fila)

    ws_stats.column_dimensions["A"].width = 35
    ws_stats.column_dimensions["B"].width = 50

    wb.save(archivo_salida)
    return ok, ocr, errores


# ============================================================
# ▶️  EJECUCIÓN PRINCIPAL
# ============================================================

def main():
    print("=" * 60)
    print("  PROCESADOR MASIVO DE PDFs → EXCEL")
    print("=" * 60)

    # Verificar carpeta
    if not os.path.exists(CARPETA_PDFS):
        print(f"\n❌ ERROR: No se encontró la carpeta: {CARPETA_PDFS}")
        print("   Edita la variable CARPETA_PDFS en el script.")
        return

    # Buscar todos los PDFs
    pdfs = [
        os.path.join(CARPETA_PDFS, f)
        for f in os.listdir(CARPETA_PDFS)
        if f.lower().endswith(".pdf")
    ]

    if not pdfs:
        print(f"\n⚠️  No se encontraron PDFs en: {CARPETA_PDFS}")
        return

    print(f"\n📁 Carpeta: {CARPETA_PDFS}")
    print(f"📄 PDFs encontrados: {len(pdfs)}")
    print(f"📊 Archivo de salida: {ARCHIVO_SALIDA}")
    print("\nIniciando procesamiento...\n")

    # Procesar PDFs con barra de progreso
    resultados = []
    for ruta in tqdm(pdfs, desc="Procesando PDFs", unit="pdf"):
        resultado = procesar_pdf(ruta)
        resultados.append(resultado)

    # Generar Excel
    print("\n📊 Generando Excel consolidado...")
    ok, ocr, errores = crear_excel(resultados, ARCHIVO_SALIDA)

    # Resumen final
    print("\n" + "=" * 60)
    print("  ✅ PROCESO COMPLETADO")
    print("=" * 60)
    print(f"  Total procesados : {len(resultados)}")
    print(f"  Correctos        : {ok}")
    print(f"  Con OCR          : {ocr}")
    print(f"  Con errores      : {errores}")
    print(f"\n  📁 Resultado guardado en: {ARCHIVO_SALIDA}")
    print("=" * 60)


if __name__ == "__main__":
    main()